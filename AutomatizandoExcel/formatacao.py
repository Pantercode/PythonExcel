from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl import Workbook
font = Font(name='Calibri', size=11,
            bold=False, 
            italic=False, 
            underline='none', 
            strike=False,
            color='FF000000')


fill = PatternFill(fill_type=None, 
                   start_color="FFFFFF", 
                   end_color="FF000000")


border = Border(
  left=Side(border_style='dashed', color='FFFF0000'),
  right=Side(border_style='dashed', color='FF000000'),
  top=Side(border_style=None, color='FF000000'),
  bottom=Side(border_style=None, color='FF000000'),
  diagonal=Side(border_style=None, color='FF000000'),  # Remove diagonal_direction argument
  outline=Side(border_style=None, color='FF000000'),
  vertical=Side(border_style=None, color='FF000000'),
  horizontal=Side(border_style=None, color='FF000000')
)

alignment = Alignment(horizontal='general',
                       vertical='bottom',
                       text_rotation=0,
                       wrap_text=False,
                       shrink_to_fit=False,
                       indent=0)
number_format = 'General'

proctection = Protection(locked=True,
                          hidden=False)

wb = Workbook()
sheet = wb.active

sheet["A1"] = "teste"
sheet["A2"] = "teste2"
sheet["A3"] = "teste3"
sheet["A4"] = "teste4"
sheet["A5"] = "teste5"

sheet["A1"].font = Font(bold=True, size=20)
sheet["A2"].font = Font(italic=True, color="FFFF0000")
sheet["A3"].border = border
sheet["A4"].alignment  = alignment 
sheet["A5"].protection = proctection
wb.save("formatacao.xlsx")