from openpyxl import load_workbook
from openpyxl.drawing.image import Image

# Carregar a planilha
wb = load_workbook(r'C:\Users\marcell.oliveira\Desktop\pythoexcel\PythonExcel\AutomatizandoExcel\exemplo.xlsx')

# Obter os nomes das planilhas
print(wb.sheetnames)

# Selecionar a planilha pelo nome
sheet = wb["Sheet1"]

# Acessar valores das células
print(sheet["A3"].value)  # Valor da célula A3
print(sheet["B2"].value)  # Valor da célula B2
print(sheet.cell(row=2, column=2).value)  # Valor da célula na linha 2, coluna 2
print(sheet.max_row)
print(sheet.max_column)
for c in range(0, sheet.max_row):
    print(sheet.cell(row=c+1, column=2).value)


for c in range(0, sheet.max_row):
    print(sheet.cell(row=2, column=c+1).value)

sheet.cell(row=2, column=3).value = 75
wb.save('exemplo2.xlsx')

#agrupamento de celulas

sheet.merge_cells('A1:D1')
sheet.cell(row=1, column=1).value = 'Agrupamento'
sheet.unmerge_cells('A1:D1')
wb.save('exemplo2.xlsx')

#deletando colunas
sheet.insert_rows(4)
sheet.delete_rows(4)
sheet.delete_cols(2, 5)
wb.save('exemplo3.xlsx')


#inserindo imagem

img = Image(r'C:\Users\marcell.oliveira\Desktop\pythoexcel\PythonExcel\AutomatizandoExcel\catlogo.png')
sheet.add_image(img, 'A2')
wb.save('exemplo4.xlsx')

#Criando uma folha nova